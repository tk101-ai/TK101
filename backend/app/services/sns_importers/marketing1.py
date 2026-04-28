"""마케팅1팀 SNS DB 엑셀 → DB importer.

엑셀 파일: `[TK내부] 2026 서울시 글로벌 SNS DB 관리.xlsx`
시트 9개를 파싱하여 SocialAccount / SocialWeeklySnapshot / SocialPost 로 적재한다.

- `parse_workbook(source)`     : 엑셀 bytes 또는 경로 → ParsedWorkbook (DB 비의존)
- `import_to_db(db, parsed)`   : ParsedWorkbook → DB 멱등 적재 (커밋은 호출자 책임)

설계 원칙
- 컬럼 위치 하드코딩 금지: 콘텐츠 시트는 행3 헤더 텍스트로 동적 매핑 (시트마다 컬럼 순서 다름)
- 멱등성:
    * 계정 (platform, language) UNIQUE
    * 스냅샷 (account_id, year, month, week_number) UNIQUE → followers UPDATE
    * 포스트는 (account_id, posted_at, title, url) 동일하면 UPDATE, 아니면 INSERT
- 데이터 무결성: 2025/2026 연도 혼재 행은 그대로 두고 보정하지 않음
"""

from __future__ import annotations

import io
import logging
from datetime import date, datetime
from typing import Any, TypedDict

import openpyxl
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.sns import SocialAccount, SocialPost, SocialWeeklySnapshot

logger = logging.getLogger(__name__)


# ------------------------------------------------------------------
# 타입 정의
# ------------------------------------------------------------------


class ParsedAccount(TypedDict):
    platform: str
    language: str
    handle: str | None


class ParsedSnapshot(TypedDict):
    platform: str
    language: str
    year: int
    month: int
    week_number: int
    followers: int


class ParsedPost(TypedDict):
    platform: str
    language: str
    posted_at: date
    title: str
    content_type: str | None
    producer: str | None
    view_count: int | None
    reach_count: int | None
    comment_count: int | None
    like_count: int | None
    share_count: int | None
    save_count: int | None
    repost_count: int | None
    total_engagement: int | None
    url: str | None
    data_recorded_at: date | None


class ParsedWorkbook(TypedDict):
    accounts: list[ParsedAccount]
    snapshots: list[ParsedSnapshot]
    posts: list[ParsedPost]


# ------------------------------------------------------------------
# 매핑 상수
# ------------------------------------------------------------------


# 콘텐츠 시트명 → (platform, language)
CONTENT_SHEET_MAP: dict[str, tuple[str, str]] = {
    "콘텐츠 DB_페이스북(영)": ("facebook", "en"),
    "콘텐츠 DB_인스타그램(영)": ("instagram", "en"),
    "콘텐츠 DB_웨이보(중간)": ("weibo", "zh"),
    "콘텐츠 DB_페이스북(중번)": ("facebook", "zh"),
    "콘텐츠 DB_페이스북(일)": ("facebook", "ja"),
    "콘텐츠 DB_인스타그램(일)": ("instagram", "ja"),
    "콘텐츠 DB_트위터(영)": ("twitter", "en"),
    "콘텐츠 DB_유튜브(영)": ("youtube", "en"),
}

# 통합 시트의 어권/플랫폼 텍스트 매핑
LANGUAGE_TEXT_MAP: dict[str, str] = {"영문": "en", "중문": "zh", "일문": "ja"}
PLATFORM_TEXT_MAP: dict[str, str] = {
    "페이스북": "facebook",
    "인스타": "instagram",
    "인스타그램": "instagram",
    "트위터": "twitter",
    "유튜브": "youtube",
    "웨이보": "weibo",
}

# 통합 시트 — 각 주차의 팔로워 컬럼 인덱스 (1-based)
WEEK_FOLLOWER_COLUMNS: dict[int, int] = {1: 4, 2: 6, 3: 8, 4: 10}

# 콘텐츠 형태 한국어 → enum
CONTENT_TYPE_MAP: dict[str, str] = {
    "이미지": "image",
    "영상": "video",
    "동영상": "video",
    "숏폼": "short",
    "쇼츠": "short",
    "릴스": "short",
}

# 콘텐츠 시트 헤더 키 (스크랩(收藏) 처럼 변형 있음 → 부분일치 키)
HEADER_ALIASES: dict[str, list[str]] = {
    "month": ["월"],
    "week": ["주차"],
    "no": ["번호"],
    "posted_at": ["배포일"],
    "title": ["콘텐츠 제목"],
    "content_type": ["콘텐츠 형태"],
    "producer": ["제작"],
    "view_count": ["조회수"],
    "reach_count": ["도달"],
    "comment_count": ["댓글"],
    "like_count": ["좋아요"],
    "share_count": ["공유"],
    "save_count": ["스크랩"],          # "스크랩(收藏)" 등 변형 부분일치
    "repost_count": ["리포스팅"],
    "total_engagement": ["토탈 인게이지먼트"],
    "url": ["링크"],
    "data_recorded_at": ["데이터 기입"],  # "데이터 기입일자" / "데이터 기입 일자"
}


# ------------------------------------------------------------------
# 헬퍼
# ------------------------------------------------------------------


def _to_int(value: Any) -> int | None:
    """엑셀 셀 → int. None / '-' / 빈문자 → None. float은 int 변환."""
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if stripped in {"", "-", "--", "N/A", "n/a"}:
            return None
        try:
            return int(float(stripped))
        except ValueError:
            return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        try:
            return int(value)
        except (ValueError, OverflowError):
            return None
    return None


def _to_date(value: Any) -> date | None:
    """엑셀 셀 → date. datetime → date, '2026.03.04' → date, 그 외 None."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d"):
            try:
                return datetime.strptime(stripped, fmt).date()
            except ValueError:
                continue
    return None


def _normalize_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _build_header_map(ws: openpyxl.worksheet.worksheet.Worksheet, header_row: int = 3) -> dict[str, int]:
    """행3 헤더를 스캔하여 alias key → column index (1-based) 사전 반환."""
    raw: dict[int, str] = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val is None:
            continue
        raw[col] = str(val).strip()

    mapped: dict[str, int] = {}
    for key, aliases in HEADER_ALIASES.items():
        for col, header in raw.items():
            if any(alias in header for alias in aliases):
                mapped[key] = col
                break
    return mapped


# ------------------------------------------------------------------
# 통합 시트 파싱
# ------------------------------------------------------------------


def _parse_summary_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    default_year: int = 2026,
) -> tuple[list[ParsedAccount], list[ParsedSnapshot]]:
    """통합 시트 → (accounts, snapshots).

    스캔 규칙:
    - 컬럼2(B)에서 "1월"~"12월" 발견 → 현재 월 갱신
    - 컬럼2에서 "영문"/"중문"/"일문" 발견 → 현재 어권 갱신, 같은 행의 컬럼3 플랫폼도 처리
    - 컬럼3(C)에서 플랫폼 텍스트 발견 → 행 시작
    - "합계" 행 / "*" 주석 행 / 헤더 행 (구분/팔로워수/첫째주 등) 스킵
    """
    accounts: dict[tuple[str, str], ParsedAccount] = {}
    snapshots: list[ParsedSnapshot] = []

    current_month: int | None = None
    current_language: str | None = None

    for row in range(1, ws.max_row + 1):
        col_b = _normalize_text(ws.cell(row=row, column=2).value)
        col_c = _normalize_text(ws.cell(row=row, column=3).value)

        # 월 표시 갱신
        if col_b and col_b.endswith("월") and len(col_b) <= 3:
            try:
                current_month = int(col_b.rstrip("월"))
                current_language = None
                continue
            except ValueError:
                pass

        # 합계 / 주석 / 헤더 행 스킵
        if col_b in {"합계", "구분"} or (col_b and col_b.startswith("*")):
            continue
        if col_b in {"첫째주", "둘째주", "셋째주", "넷째주"}:
            continue

        # 어권 갱신 (해당 행에 플랫폼이 동시에 있을 수 있음)
        if col_b in LANGUAGE_TEXT_MAP:
            current_language = LANGUAGE_TEXT_MAP[col_b]

        # 플랫폼 행: 어권 + 월이 모두 있어야 의미 있음
        if not col_c or current_language is None or current_month is None:
            continue
        platform = PLATFORM_TEXT_MAP.get(col_c)
        if platform is None:
            continue

        key = (platform, current_language)
        if key not in accounts:
            accounts[key] = ParsedAccount(
                platform=platform,
                language=current_language,
                handle=None,
            )

        # 4개 주차 팔로워
        for week_number, col in WEEK_FOLLOWER_COLUMNS.items():
            followers = _to_int(ws.cell(row=row, column=col).value)
            if followers is None:
                continue
            snapshots.append(
                ParsedSnapshot(
                    platform=platform,
                    language=current_language,
                    year=default_year,
                    month=current_month,
                    week_number=week_number,
                    followers=followers,
                )
            )

    return list(accounts.values()), snapshots


# ------------------------------------------------------------------
# 콘텐츠 시트 파싱
# ------------------------------------------------------------------


def _parse_content_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    platform: str,
    language: str,
) -> list[ParsedPost]:
    """콘텐츠 DB 시트 → posts. 행3 헤더 기반 동적 매핑."""
    header_map = _build_header_map(ws, header_row=3)

    if "title" not in header_map or "posted_at" not in header_map:
        logger.warning("필수 헤더 누락 — sheet=%s headers=%s", ws.title, header_map)
        return []

    posts: list[ParsedPost] = []
    for row in range(4, ws.max_row + 1):
        title = _normalize_text(ws.cell(row=row, column=header_map["title"]).value)
        if not title:
            continue

        posted_at = _to_date(ws.cell(row=row, column=header_map["posted_at"]).value)
        if posted_at is None:
            # 배포일 없는 행은 건너뜀 (PK 결정 불가)
            continue

        raw_type = _normalize_text(ws.cell(row=row, column=header_map["content_type"]).value) if "content_type" in header_map else None
        content_type = CONTENT_TYPE_MAP.get(raw_type) if raw_type else None

        producer = _normalize_text(ws.cell(row=row, column=header_map["producer"]).value) if "producer" in header_map else None
        url = _normalize_text(ws.cell(row=row, column=header_map["url"]).value) if "url" in header_map else None

        def cell_int(key: str) -> int | None:
            col = header_map.get(key)
            if col is None:
                return None
            return _to_int(ws.cell(row=row, column=col).value)

        data_recorded_at = (
            _to_date(ws.cell(row=row, column=header_map["data_recorded_at"]).value)
            if "data_recorded_at" in header_map
            else None
        )

        posts.append(
            ParsedPost(
                platform=platform,
                language=language,
                posted_at=posted_at,
                title=title,
                content_type=content_type,
                producer=producer,
                view_count=cell_int("view_count"),
                reach_count=cell_int("reach_count"),
                comment_count=cell_int("comment_count"),
                like_count=cell_int("like_count"),
                share_count=cell_int("share_count"),
                save_count=cell_int("save_count"),
                repost_count=cell_int("repost_count"),
                total_engagement=cell_int("total_engagement"),
                url=url,
                data_recorded_at=data_recorded_at,
            )
        )

    return posts


# ------------------------------------------------------------------
# parse_workbook
# ------------------------------------------------------------------


def parse_workbook(source: bytes | str, default_year: int = 2026) -> ParsedWorkbook:
    """엑셀 bytes 또는 경로를 받아 파싱.

    Args:
        source: bytes(업로드 파일) 또는 str(파일 경로).
        default_year: 통합 시트에 연도가 없을 때 기본값 (2026).
    """
    if isinstance(source, bytes):
        wb = openpyxl.load_workbook(io.BytesIO(source), data_only=True, read_only=False)
    else:
        wb = openpyxl.load_workbook(source, data_only=True, read_only=False)

    accounts: list[ParsedAccount] = []
    snapshots: list[ParsedSnapshot] = []
    posts: list[ParsedPost] = []

    # 통합 시트
    if "통합" in wb.sheetnames:
        accounts, snapshots = _parse_summary_sheet(wb["통합"], default_year=default_year)

    # 콘텐츠 시트로부터도 계정 보강 (통합에 없는 조합 안전망)
    account_keys = {(a["platform"], a["language"]) for a in accounts}
    for sheet_name, (platform, language) in CONTENT_SHEET_MAP.items():
        if sheet_name not in wb.sheetnames:
            logger.warning("시트 누락: %s", sheet_name)
            continue
        if (platform, language) not in account_keys:
            accounts.append(ParsedAccount(platform=platform, language=language, handle=None))
            account_keys.add((platform, language))
        posts.extend(_parse_content_sheet(wb[sheet_name], platform, language))

    return ParsedWorkbook(accounts=accounts, snapshots=snapshots, posts=posts)


# ------------------------------------------------------------------
# import_to_db (멱등)
# ------------------------------------------------------------------


async def _ensure_accounts(
    db: AsyncSession, parsed_accounts: list[ParsedAccount]
) -> tuple[dict[tuple[str, str], Any], int]:
    """계정을 (platform, language) 키로 멱등 보장. 반환: (key→id 매핑, 신규 추가 수)."""
    added = 0
    key_to_id: dict[tuple[str, str], Any] = {}

    existing_rows = (await db.execute(select(SocialAccount))).scalars().all()
    for acc in existing_rows:
        key_to_id[(acc.platform, acc.language)] = acc.id

    for parsed in parsed_accounts:
        key = (parsed["platform"], parsed["language"])
        if key in key_to_id:
            continue
        new_acc = SocialAccount(
            platform=parsed["platform"],
            language=parsed["language"],
            handle=parsed.get("handle"),
            is_active=True,
        )
        db.add(new_acc)
        await db.flush()  # id 확보
        key_to_id[key] = new_acc.id
        added += 1

    return key_to_id, added


async def _upsert_snapshots(
    db: AsyncSession,
    parsed_snapshots: list[ParsedSnapshot],
    key_to_id: dict[tuple[str, str], Any],
) -> tuple[int, int]:
    added = 0
    updated = 0
    for snap in parsed_snapshots:
        account_id = key_to_id.get((snap["platform"], snap["language"]))
        if account_id is None:
            continue
        result = await db.execute(
            select(SocialWeeklySnapshot).where(
                SocialWeeklySnapshot.account_id == account_id,
                SocialWeeklySnapshot.year == snap["year"],
                SocialWeeklySnapshot.month == snap["month"],
                SocialWeeklySnapshot.week_number == snap["week_number"],
            )
        )
        existing = result.scalar_one_or_none()
        if existing is None:
            db.add(
                SocialWeeklySnapshot(
                    account_id=account_id,
                    year=snap["year"],
                    month=snap["month"],
                    week_number=snap["week_number"],
                    followers=snap["followers"],
                )
            )
            added += 1
        elif existing.followers != snap["followers"]:
            existing.followers = snap["followers"]
            updated += 1
    return added, updated


async def _upsert_posts(
    db: AsyncSession,
    parsed_posts: list[ParsedPost],
    key_to_id: dict[tuple[str, str], Any],
) -> tuple[int, int]:
    added = 0
    updated = 0
    post_fields = (
        "content_type",
        "producer",
        "view_count",
        "reach_count",
        "comment_count",
        "like_count",
        "share_count",
        "save_count",
        "repost_count",
        "total_engagement",
        "data_recorded_at",
    )
    for post in parsed_posts:
        account_id = key_to_id.get((post["platform"], post["language"]))
        if account_id is None:
            continue
        # dedup: (account_id, posted_at, title, url) 동일하면 UPDATE
        result = await db.execute(
            select(SocialPost).where(
                SocialPost.account_id == account_id,
                SocialPost.posted_at == post["posted_at"],
                SocialPost.title == post["title"],
                SocialPost.url == post["url"],
            )
        )
        existing = result.scalar_one_or_none()
        if existing is None:
            db.add(
                SocialPost(
                    account_id=account_id,
                    posted_at=post["posted_at"],
                    title=post["title"],
                    content_type=post["content_type"],
                    producer=post["producer"],
                    view_count=post["view_count"],
                    reach_count=post["reach_count"],
                    comment_count=post["comment_count"],
                    like_count=post["like_count"],
                    share_count=post["share_count"],
                    save_count=post["save_count"],
                    repost_count=post["repost_count"],
                    total_engagement=post["total_engagement"],
                    url=post["url"],
                    data_recorded_at=post["data_recorded_at"],
                )
            )
            added += 1
        else:
            mutated = False
            for field in post_fields:
                new_val = post[field]  # type: ignore[literal-required]
                if new_val is None:
                    continue
                if getattr(existing, field) != new_val:
                    setattr(existing, field, new_val)
                    mutated = True
            if mutated:
                updated += 1
    return added, updated


async def import_to_db(
    db: AsyncSession,
    parsed: ParsedWorkbook,
    default_year: int = 2026,
) -> dict[str, int]:
    """파싱 결과를 DB에 멱등 적재.

    Args:
        db: AsyncSession (커밋은 호출자 책임).
        parsed: parse_workbook 결과.
        default_year: ParsedSnapshot에 year가 없는 경우 기본값 (현재는 항상 채워짐).

    Returns:
        {"accounts_added": N, "snapshots_added": N, "snapshots_updated": N,
         "posts_added": N, "posts_updated": N}
    """
    key_to_id, accounts_added = await _ensure_accounts(db, parsed["accounts"])
    snapshots_added, snapshots_updated = await _upsert_snapshots(db, parsed["snapshots"], key_to_id)
    posts_added, posts_updated = await _upsert_posts(db, parsed["posts"], key_to_id)

    await db.flush()

    return {
        "accounts_added": accounts_added,
        "snapshots_added": snapshots_added,
        "snapshots_updated": snapshots_updated,
        "posts_added": posts_added,
        "posts_updated": posts_updated,
    }
