"""종합관리시트 엑셀 파서 (T9 Phase B-1).

대상 시트: "래더엑스_종합관리시트" 형태의 pivoted 매트릭스.
- 헤더 행: 회사 / 기간 / 1201_1207 / 1208_1214 / ... / 월계 / 누계
- 데이터 행: 회사명 / 항목명(KR_매입, VN_재고이동 등) / 주차별 금액

처리 전략:
1. 시트 전체 스캔하며 "회사" + "기간" 같이 있는 헤더 행 탐지 (여러 회계연도 지원).
2. 헤더에서 period 컬럼 (정규식 ``\\d{4}_\\d{4}``) 인덱스 추출. 월계/누계/1분기계 등 집계는 skip.
3. 다음 행부터 빈 행 만날 때까지 항목 행으로 처리.
4. 회사명은 첫 데이터 행에서 추출 (이후 행은 빈 셀이라 전파).
5. 항목별로 모은 row dict → (회사, 기간) 단위로 재조합하여 list[WeeklyRow] 반환.

period_label "1201_1207" → period_start/end 변환:
- 년도는 헤더 행 근처의 "2026" 같은 마커에서 추출 (없으면 imported_at 연도로 폴백).
- 월일 4자리는 MM/DD 로 해석.
"""
from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import openpyxl

logger = logging.getLogger(__name__)

# 주차 컬럼 정규식: "1201_1207", "0101_0107" 등.
_PERIOD_RX = re.compile(r"^(\d{2})(\d{2})_(\d{2})(\d{2})$")
# 항목명 패턴 → DB 컬럼 매핑. 줄바꿈/공백/괄호 무시.
_ITEM_MAP: dict[str, str] = {
    "KR_매입": "kr_purchase",
    "VN_재고이동": "vn_inventory_move",
    "VN_매출완료": "vn_sales_completed",
    "KR_매입 입금요청": "kr_purchase_deposit_req",
    "VN_재고이동 입금요청": "vn_inventory_deposit_req",
    "VN_매출 입금요청": "vn_sales_deposit_req",
    "계좌입금": "account_deposit",
    "현금": "cash_deposit",
}


@dataclass
class WeeklyRow:
    """파싱된 1주차 데이터 1행. 라우터가 DB upsert 시 사용."""

    company_label: str
    period_label: str
    period_start: date
    period_end: date
    values: dict[str, Decimal] = field(default_factory=dict)
    raw_row: dict[str, Any] = field(default_factory=dict)


@dataclass
class SummaryParseResult:
    """파서 1회 실행 결과."""

    rows: list[WeeklyRow]
    warnings: list[str]


def _normalize_item_name(raw: object) -> str | None:
    """엑셀 셀의 항목명을 정규화. 줄바꿈·공백·괄호 안 내용 제거."""
    if raw is None:
        return None
    s = str(raw)
    # 괄호 안 부연설명 제거 (예: "KR_매입\n ( vat 미포함)" → "KR_매입")
    s = re.sub(r"[\(（][^)）]*[\)）]", "", s)
    # 줄바꿈/연속공백 정리
    s = re.sub(r"\s+", " ", s).strip()
    return s or None


def _to_decimal(value: object) -> Decimal | None:
    """엑셀 셀 → Decimal. 숫자 아니면 None."""
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def _detect_periods(header_row: tuple[object, ...]) -> list[tuple[int, str]]:
    """헤더 행에서 (column_index, period_label) 페어 추출.

    "1201_1207" 패턴만 통과. "월계", "누계", "1분기 계" 등은 skip.
    """
    out: list[tuple[int, str]] = []
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        label = str(cell).strip()
        if _PERIOD_RX.match(label):
            out.append((idx, label))
    return out


def _detect_year_marker(rows: list[tuple[object, ...]], header_row_idx: int) -> int | None:
    """헤더 위쪽 5행 내에서 "2026" 같은 연도 마커 찾기.

    종합관리시트는 회계연도마다 헤더가 따로 있고 그 위에 연도 표기.
    찾지 못하면 None 반환 (caller 가 imported_at 연도로 폴백).
    """
    start = max(0, header_row_idx - 5)
    for row in rows[start:header_row_idx]:
        for cell in row:
            if cell is None:
                continue
            s = str(cell).strip()
            if re.fullmatch(r"20\d{2}", s):
                return int(s)
    return None


def _period_label_to_dates(
    label: str, year: int
) -> tuple[date, date] | None:
    """``1201_1207`` + year=2025 → (2025-12-01, 2025-12-07)."""
    m = _PERIOD_RX.match(label)
    if not m:
        return None
    sm, sd, em, ed = m.groups()
    try:
        start = date(year, int(sm), int(sd))
        end = date(year, int(em), int(ed))
    except ValueError:
        return None
    # 연말 → 연초 전환 케이스 (예: 1229_0104). 종료가 시작보다 작으면 연도 + 1.
    if end < start:
        try:
            end = date(year + 1, int(em), int(ed))
        except ValueError:
            return None
    return start, end


def parse_summary_sheet(
    path: Path | str,
    *,
    sheet_name: str | None = None,
    fallback_year: int | None = None,
) -> SummaryParseResult:
    """종합관리시트 엑셀 파일을 파싱하여 주차별 행 리스트 반환.

    sheet_name: None 이면 '래더엑스_종합관리시트' 또는 첫 번째 시트 사용.
    fallback_year: 연도 마커 못 찾으면 사용할 연도. None 이면 오늘 연도.
    """
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    if sheet_name is None:
        sheet_name = (
            "래더엑스_종합관리시트"
            if "래더엑스_종합관리시트" in wb.sheetnames
            else wb.sheetnames[0]
        )
    ws = wb[sheet_name]

    rows: list[tuple[object, ...]] = list(ws.iter_rows(values_only=True))
    fallback_year = fallback_year or datetime.now().year

    parsed: list[WeeklyRow] = []
    warnings: list[str] = []

    row_idx = 0
    while row_idx < len(rows):
        row = rows[row_idx]
        # 헤더 행 탐지: "회사" + "기간" 셀 동시 존재 (인접 컬럼).
        cell_values = [str(c).strip() if c is not None else "" for c in row]
        if "회사" in cell_values and "기간" in cell_values:
            # 헤더 발견.
            periods = _detect_periods(row)
            if not periods:
                row_idx += 1
                continue
            year = _detect_year_marker(rows, row_idx) or fallback_year

            # 회사명: 헤더 바로 아래 행의 "회사" 컬럼.
            company_col = cell_values.index("회사")
            item_col = cell_values.index("기간")

            # 헤더 다음 행부터 항목 행 수집.
            data_start = row_idx + 1
            company_label: str | None = None
            # period 단위로 누적할 dict: {period_label: {item_col: value}}
            per_period: dict[str, dict[str, Decimal]] = {
                lbl: {} for _, lbl in periods
            }
            per_period_raw: dict[str, dict[str, Any]] = {
                lbl: {} for _, lbl in periods
            }

            scan = data_start
            while scan < len(rows):
                drow = rows[scan]
                # 전체 빈 행이면 블록 종료.
                if all(c is None or str(c).strip() == "" for c in drow):
                    scan += 1
                    break
                # 다음 헤더 만나면 종료.
                drow_vals = [
                    str(c).strip() if c is not None else "" for c in drow
                ]
                if "회사" in drow_vals and "기간" in drow_vals:
                    break

                # 회사명 갱신 (첫 행에만 있고 이후엔 빈 셀이 많음).
                if drow[company_col] is not None:
                    candidate = str(drow[company_col]).strip()
                    if candidate and candidate != "회사":
                        company_label = candidate

                item_raw = drow[item_col] if item_col < len(drow) else None
                item = _normalize_item_name(item_raw)
                if item and item in _ITEM_MAP:
                    col_key = _ITEM_MAP[item]
                    for idx, period_label in periods:
                        v = _to_decimal(drow[idx] if idx < len(drow) else None)
                        if v is not None:
                            per_period[period_label][col_key] = v
                            per_period_raw[period_label][item] = str(
                                drow[idx]
                            )
                scan += 1

            # per_period 를 WeeklyRow 로 변환.
            if not company_label:
                warnings.append(
                    f"row {row_idx + 1}: 회사명 추출 실패 — 건너뜀"
                )
                row_idx = scan
                continue

            for period_label, values in per_period.items():
                if not values:
                    continue  # 빈 주차는 skip
                dates = _period_label_to_dates(period_label, year)
                if dates is None:
                    warnings.append(
                        f"period_label '{period_label}' 날짜 변환 실패"
                    )
                    continue
                parsed.append(
                    WeeklyRow(
                        company_label=company_label,
                        period_label=period_label,
                        period_start=dates[0],
                        period_end=dates[1],
                        values=values,
                        raw_row=per_period_raw[period_label],
                    )
                )

            row_idx = scan
        else:
            row_idx += 1

    return SummaryParseResult(rows=parsed, warnings=warnings)
