"""면장(통관신고) 파서 (Priority 4) — 엑셀 + PDF.

products_parser.py 와 동일 패턴:
- openpyxl(.xlsx/.xlsm) / pdfplumber·pdfminer(.pdf), **헤더 텍스트 기반** 컬럼 매핑
  (위치 기반 X — 양식 변경에 강함).
- dataclass 반환, 원본 행(raw_row) 보존.
- 엑셀·PDF 가 동일한 ``_HEADER_CANDIDATES`` / ``reverse_calc_actual_price`` /
  ``CustomsRow`` 로 수렴한다. 진입점은 확장자로 분기하는 ``parse_customs_file``.

핵심 비즈니스 규칙 (역산):
- 면장의 신고가(declared_price)는 관세 절감 목적으로 실제 가치의 75% 로 신고된다.
- 실가 역산: ``actual_price = round(declared_price / ratio, 2)``.
  ratio 기본값은 config.py 의 ``distribution_customs_declare_ratio`` (= 0.75).
- ratio <= 0 또는 declared_price 누락/0 이면 actual_price 는 None (안전 가드).

수집 컬럼:
- declaration_number (신고번호/면장번호)
- declared_price (신고가/단가/가격)
- stock_qty (재고/재고수량/수량)
- product (품명/상품명)
- bl_number (BL번호)
- currency (통화)
- declared_at (신고일자)

⚠️ 실제 면장 엑셀/PDF 샘플이 아직 없다. 아래 ``_HEADER_CANDIDATES`` 의 한국어
헤더 후보(엑셀·PDF 표 공용)와 ``parse_customs_pdf`` 섹션의 정규식
(``_PDF_DECL_NO_PATTERN`` 등)은 모두 추정치이며, 샘플 도착 시 그 부분만
수정하면 된다.
"""
from __future__ import annotations

import io
import logging
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import openpyxl

from app.config import settings

logger = logging.getLogger(__name__)

# PDF 파싱 가드. 면장 PDF 는 통상 1~몇 페이지. 텍스트 레이어 없는(스캔) PDF 가
# 수백 페이지일 경우 pdfplumber 레이아웃 분석이 메모리/CPU 를 폭증시킬 수 있어
# 페이지 수를 제한한다 (zip-bomb/huge-page DoS 완화).
_PDF_MAX_PAGES = 50


# ===========================================================================
# 컬럼 매핑 — 실제 면장 엑셀 샘플 도착 시 여기만 수정하면 된다.
# TODO: confirm against real 면장 Excel — 헤더 텍스트는 모두 추정치.
# 각 DB 필드별로 가능한 한국어 헤더 후보 리스트. 정규화된 헤더가 후보에 있으면 매핑.
# ===========================================================================
_HEADER_CANDIDATES: dict[str, list[str]] = {
    # 신고번호(declaration number) — 면장 식별자. partial UNIQUE 키.
    "declaration_number": ["신고번호", "면장번호", "수입신고번호", "통관번호"],
    # 신고가(declared price) — 실가의 75% 로 신고된 금액.
    "declared_price": ["단가", "가격", "신고가격", "신고가", "신고금액"],
    # 재고(stock quantity).
    "stock_qty": ["재고", "재고수량", "수량"],
    # 품명(product name).
    "product": ["품명", "상품명", "제품명", "품목"],
    # 연결 BL 번호.
    "bl_number": ["BL번호", "B/L번호", "BL No", "비엘번호"],
    # 통화.
    "currency": ["통화", "화폐", "currency"],
    # 신고일자.
    "declared_at": ["신고일자", "신고일", "통관일자", "수입신고일"],
}

# 헤더 행으로 인정하려면 최소한 이 필드들 중 하나는 발견돼야 한다.
# (신고번호 또는 신고가 — 면장의 핵심 식별/금액 컬럼.)
_REQUIRED_ANY: tuple[str, ...] = ("declaration_number", "declared_price")


@dataclass
class CustomsRow:
    """파싱된 면장 1행. actual_price 는 파서/서비스가 역산해 채운다.

    한국 관세청 신고필증 항목 번호 매핑은 모델 docstring 참조.
    declaration_type 이 'export' 면 75% 역산 미적용, 'import' 또는 NULL 이면 적용.
    """

    declaration_number: str | None = None
    # "export" / "import" / None (미지).
    declaration_type: str | None = None
    item_name: str | None = None  # ㉗ 품명 (예: "HAND BAG")
    product: str | None = None  # ㉚ 모델·규격 (예: "GUCCI BAG")
    bl_number: str | None = None
    unit_price: Decimal | None = None  # ㉝ 단가
    declared_price: Decimal | None = None  # ㊳ 신고가격(FOB) — 통상 USD
    declared_price_krw: Decimal | None = None  # ㊳ 옆 KRW 환산
    # actual_price: 수입은 declared / ratio, 수출은 declared 그대로 (서비스 계산).
    actual_price: Decimal | None = None
    currency: str | None = None
    stock_qty: int | None = None
    declared_at: date | None = None
    raw_row: dict[str, Any] = field(default_factory=dict)


@dataclass
class CustomsParseResult:
    rows: list[CustomsRow]
    warnings: list[str]


def _normalize_header(raw: object) -> str | None:
    if raw is None:
        return None
    # 병합셀/주석 줄바꿈은 첫 줄만 사용 (예: "신고번호\n(Customs No.)" → "신고번호").
    # 주의: 공백 정규화를 먼저 하면 \n 이 사라져 첫줄 추출이 무력화되므로 split 을 먼저.
    first_line = str(raw).split("\n", 1)[0]
    s = re.sub(r"\s+", " ", first_line).strip()
    return s or None


def _to_int(value: object) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(Decimal(str(value)))
    except (InvalidOperation, ValueError):
        return None


def _to_decimal(value: object) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        # "1,000,000" 같은 천단위 콤마 제거.
        cleaned = str(value).replace(",", "").strip()
        if cleaned == "":
            return None
        return Decimal(cleaned)
    except (InvalidOperation, ValueError):
        return None


def _to_date(value: object) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in ("%Y.%m.%d", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def reverse_calc_actual_price(
    declared_price: Decimal | None,
    *,
    declaration_type: str | None = None,
    ratio: float | None = None,
) -> Decimal | None:
    """신고가 → 실가 역산. declaration_type 에 따라 분기.

    - declaration_type == 'export' (수출신고필증):
        신고가격(FOB)이 실가에 가까움 → actual_price = declared_price (역산 미적용).
        의미 없는 유령 수치(예: 9,834/0.75 = 13,112) 방지.
    - declaration_type == 'import' 또는 None (수입/미지):
        ``actual_price = round(declared_price / ratio, 2)``. 관세 절감 목적으로
        실가의 75%로 신고된 가정. (legacy 행 하위호환을 위해 NULL 도 import 로 취급.)

    안전 가드:
    - declared_price 가 None 또는 0 이하 → None (역산 불가/무의미).
    - ratio 가 None 이면 config.distribution_customs_declare_ratio 사용.
    - ratio <= 0 → None (0 나눗셈/음수 비율 방지).
    """
    if declared_price is None:
        return None
    if declared_price <= 0:
        return None
    if declaration_type == "export":
        # 수출은 역산 미적용 — 화면 일관성을 위해 declared 값 그대로 노출.
        return declared_price.quantize(Decimal("0.01"))
    effective_ratio = (
        ratio if ratio is not None else settings.distribution_customs_declare_ratio
    )
    if effective_ratio <= 0:
        logger.warning(
            "면장 역산 비율이 0 이하 (%s) — actual_price 계산 생략", effective_ratio
        )
        return None
    return (declared_price / Decimal(str(effective_ratio))).quantize(Decimal("0.01"))


def _build_col_map(row: tuple[object, ...]) -> dict[str, int]:
    """엑셀 1행에서 {DB 필드: 컬럼 인덱스} 매핑 추출 (헤더 후보 매칭)."""
    col_map: dict[str, int] = {}
    for col_idx, cell in enumerate(row):
        header = _normalize_header(cell)
        if header is None:
            continue
        for field_name, candidates in _HEADER_CANDIDATES.items():
            if field_name in col_map:
                continue
            if header in candidates:
                col_map[field_name] = col_idx
                break
    return col_map


def _find_header_row(
    rows: list[tuple[object, ...]], max_scan: int = 20
) -> tuple[int, dict[str, int]] | None:
    """헤더 행 인덱스 + 컬럼 매핑 반환. 핵심 컬럼 1개 이상 있어야 헤더로 인정."""
    for idx, row in enumerate(rows[:max_scan]):
        col_map = _build_col_map(row)
        if any(key in col_map for key in _REQUIRED_ANY):
            return idx, col_map
    return None


def _row_to_customs(
    row: tuple[object, ...],
    col_map: dict[str, int],
    *,
    ratio: float,
) -> CustomsRow | None:
    """매핑된 1행(tuple) → CustomsRow. 식별 가능한 값이 없으면 None.

    엑셀·PDF 표 양쪽이 공유한다 (헤더 매핑·역산 로직 단일화).
    """

    def get(col_key: str) -> object:
        col_idx = col_map.get(col_key)
        if col_idx is None or col_idx >= len(row):
            return None
        return row[col_idx]

    declaration_number = get("declaration_number")
    product = get("product")
    declared_price = _to_decimal(get("declared_price"))

    # 식별 가능한 값이 전혀 없으면 skip (합계/소계/빈 행 등).
    if (
        (declaration_number is None or str(declaration_number).strip() == "")
        and (product is None or str(product).strip() == "")
        and declared_price is None
    ):
        return None

    return CustomsRow(
        declaration_number=(
            str(declaration_number).strip()
            if declaration_number not in (None, "")
            else None
        ),
        product=(str(product).strip() if product not in (None, "") else None),
        bl_number=(
            str(get("bl_number")).strip()
            if get("bl_number") not in (None, "")
            else None
        ),
        declared_price=declared_price,
        actual_price=reverse_calc_actual_price(declared_price, ratio=ratio),
        currency=(
            str(get("currency")).strip() if get("currency") not in (None, "") else None
        ),
        stock_qty=_to_int(get("stock_qty")),
        declared_at=_to_date(get("declared_at")),
        raw_row={
            k: (str(row[v]) if v < len(row) and row[v] is not None else None)
            for k, v in col_map.items()
        },
    )


def _rows_to_customs(
    rows: list[tuple[object, ...]],
    header_idx: int,
    col_map: dict[str, int],
) -> list[CustomsRow]:
    """헤더 행 다음부터 데이터 행을 CustomsRow 리스트로 변환 (엑셀·PDF 공용)."""
    ratio = settings.distribution_customs_declare_ratio
    parsed: list[CustomsRow] = []
    for row in rows[header_idx + 1 :]:
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        customs = _row_to_customs(row, col_map, ratio=ratio)
        if customs is not None:
            parsed.append(customs)
    return parsed


def parse_customs_sheet(
    file_bytes: bytes,
    *,
    sheet_name: str | None = None,
) -> CustomsParseResult:
    """면장 엑셀 bytes → CustomsRow 리스트.

    Args:
        file_bytes: 업로드된 .xlsx/.xlsm 의 raw bytes.
        sheet_name: 특정 시트명. None 이면 첫 번째(활성) 시트 사용.

    Returns:
        CustomsParseResult(rows, warnings). 헤더 못 찾으면 rows=[] + warning.
    """
    wb = openpyxl.load_workbook(
        io.BytesIO(file_bytes), data_only=True, read_only=True
    )
    if sheet_name is not None:
        if sheet_name not in wb.sheetnames:
            return CustomsParseResult(
                rows=[],
                warnings=[f"시트 '{sheet_name}' 없음. 가용: {wb.sheetnames}"],
            )
        ws = wb[sheet_name]
    else:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))

    detected = _find_header_row(rows)
    if detected is None:
        return CustomsParseResult(
            rows=[],
            warnings=[
                "면장 헤더 행을 찾지 못했습니다 "
                "(신고번호/신고가 등). 컬럼 매핑(_HEADER_CANDIDATES) 확인 필요."
            ],
        )
    header_idx, col_map = detected

    parsed = _rows_to_customs(rows, header_idx, col_map)
    return CustomsParseResult(rows=parsed, warnings=[])


# ===========================================================================
# PDF 파서 — 면장 PDF (.pdf).
#
# ⚠️ 실제 면장 PDF 샘플이 아직 없다. 아래 두 경로(표 추출 / 라인 기반) 모두
#    레이아웃 추정에 기반하며, 샘플 도착 시 이 섹션만 조정하면 된다.
#
# 전략 (회수율 우선, 크래시 금지):
#   1) pdfplumber 로 페이지별 표(table)를 추출 → 표 행을 엑셀 행처럼 취급해
#      _find_header_row / _rows_to_customs 재사용 (헤더·역산 로직 단일화).
#   2) 표가 없거나 헤더를 못 찾으면 pdfminer/pdfplumber 본문 텍스트를
#      라인 단위로 파싱 (정규식 기반 fallback).
#   3) 어느 쪽도 못 뽑으면 rows=[] + 경고 (예외 전파 금지).
# ===========================================================================

# TODO: confirm against real 면장 PDF layout — 라인 기반 fallback 의 신고번호 패턴.
# 면장 신고번호는 통상 "12345-12-345678" 또는 11~15자리 숫자 형태로 추정.
# 실제 샘플 확보 시 이 정규식을 면장 신고번호 실제 포맷에 맞춰 좁혀야 한다.
_PDF_DECL_NO_PATTERN = re.compile(r"\b(\d{5}-\d{2}-\d{6}|\d{11,15})\b")

# TODO: confirm against real 면장 PDF layout — 금액(신고가) 추출 패턴.
# 라인 안의 천단위 콤마 포함 금액 토큰. 여러 개면 가장 큰 값을 신고가로 추정.
_PDF_AMOUNT_PATTERN = re.compile(r"\b\d{1,3}(?:,\d{3})+(?:\.\d+)?\b|\b\d{4,}(?:\.\d+)?\b")


def _extract_pdf_tables(file_bytes: bytes) -> list[list[tuple[object, ...]]]:
    """pdfplumber 로 페이지별 표를 추출. 각 표는 행(tuple) 리스트.

    pdfplumber 미설치/추출 실패 시 빈 리스트 (caller 가 라인 fallback 으로 진행).
    페이지 수는 _PDF_MAX_PAGES 로 제한 (huge-page DoS 완화).
    """
    try:
        import pdfplumber
    except ImportError:
        logger.debug("pdfplumber 미설치 — 면장 PDF 표 추출 스킵")
        return []

    tables: list[list[tuple[object, ...]]] = []
    try:
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            for page in pdf.pages[:_PDF_MAX_PAGES]:
                for table in page.extract_tables() or []:
                    # TODO: confirm against real 면장 PDF layout — 셀 → 행 매핑.
                    rows = [tuple(cell for cell in row) for row in table if row]
                    if rows:
                        tables.append(rows)
    except Exception as exc:  # noqa: BLE001 — pdfplumber 내부 예외 종류가 다양
        logger.warning("면장 PDF 표 추출 실패: %s", exc)
        return []
    return tables


def _extract_pdf_text(file_bytes: bytes) -> str:
    """면장 PDF 본문 텍스트. pdfminer 1차, 빈 결과면 pdfplumber 2차 (extractor.py 패턴)."""
    try:
        from pdfminer.high_level import extract_text as pdf_extract

        primary = (pdf_extract(io.BytesIO(file_bytes)) or "").replace("\x00", "").strip()
        if primary:
            return primary
    except Exception as exc:  # noqa: BLE001 — pdfminer 내부 예외 종류가 다양
        logger.debug("pdfminer 추출 실패, pdfplumber 시도: %s", exc)

    try:
        import pdfplumber
    except ImportError:
        return ""
    try:
        parts: list[str] = []
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            for page in pdf.pages[:_PDF_MAX_PAGES]:
                page_text = page.extract_text() or ""
                if page_text.strip():
                    parts.append(page_text)
        return "\n".join(parts).replace("\x00", "").strip()
    except Exception as exc:  # noqa: BLE001
        logger.warning("면장 PDF 본문 추출 실패: %s", exc)
        return ""


def _parse_pdf_lines(text: str) -> list[CustomsRow]:
    """본문 텍스트 라인 fallback — 신고번호+금액 정규식으로 행 추정.

    표 추출이 실패한 PDF 용 최후 수단. 신고번호가 있는 라인만 1행으로 본다.
    금액 토큰이 여러 개면 최댓값을 신고가로 추정 (TODO: 실제 양식 확인).
    """
    ratio = settings.distribution_customs_declare_ratio
    parsed: list[CustomsRow] = []
    for line in text.splitlines():
        stripped = line.strip()
        if not stripped:
            continue
        decl_match = _PDF_DECL_NO_PATTERN.search(stripped)
        if decl_match is None:
            continue
        # 신고번호 토큰을 제거한 뒤 금액을 스캔한다 (신고번호의 숫자 조각이
        # 금액 후보로 오인되는 것을 방지).
        remainder = stripped.replace(decl_match.group(1), " ", 1)
        declared_price = _largest_amount(remainder)
        parsed.append(
            CustomsRow(
                declaration_number=decl_match.group(1),
                declared_price=declared_price,
                actual_price=reverse_calc_actual_price(declared_price, ratio=ratio),
                raw_row={"line": stripped},
            )
        )
    return parsed


def _largest_amount(line: str) -> Decimal | None:
    """라인 안의 금액 토큰 중 최댓값을 Decimal 로. 없으면 None."""
    amounts = [
        amt
        for token in _PDF_AMOUNT_PATTERN.findall(line)
        if (amt := _to_decimal(token)) is not None
    ]
    return max(amounts) if amounts else None


def parse_customs_pdf(
    file_bytes: bytes,
    *,
    source_file_name: str | None = None,
) -> CustomsParseResult:
    """면장 PDF bytes → CustomsRow 리스트.

    경로 우선순위 (양식 변경에 강한 순서):
    1차: 텍스트 추출 → LLM (Haiku) 의미 기반 추출.
         면장은 통관사·품목마다 레이아웃이 달라 헤더/정규식이 항상 깨진다.
         LLM 한 번이 양식 N 종을 동시에 처리한다.
    2차: pdfplumber 표 추출 → 헤더 매핑 (양식이 표준 헤더를 쓰는 경우).
    3차: 본문 텍스트 라인 파싱 (정규식 fallback).

    LLM 경로는 키 미설정/응답 깨짐/0건 등 어떤 사유로도 fallback 으로 넘어간다.
    어느 경로에서든 1건이라도 잡으면 그 결과를 즉시 반환.

    Returns:
        CustomsParseResult(rows, warnings). PDF 레이아웃 미확정이므로 추출
        실패 시 사용자에게 경고로 안내한다.
    """
    warnings: list[str] = []

    # ---- 1차: LLM 의미 기반 추출 -----------------------------------------
    text = _extract_pdf_text(file_bytes)
    if not text:
        # 텍스트 레이어 자체가 없으면 표/라인 추출도 무의미 — 바로 종료.
        return CustomsParseResult(
            rows=[],
            warnings=[
                "면장 PDF 에서 텍스트를 추출하지 못했습니다 "
                "(스캔 이미지 PDF 이거나 텍스트 레이어 없음). "
                "텍스트 기반 PDF 인지 확인하세요."
            ],
        )

    # 순환 import 방지 — extractor 가 이 모듈의 CustomsRow / reverse_calc 를 사용한다.
    from app.services.distribution.customs_llm_extractor import (  # noqa: PLC0415
        extract_customs_from_text,
    )

    llm_result = extract_customs_from_text(
        text, source_file_name=source_file_name
    )
    if llm_result.rows:
        # 성공 — LLM 결과를 그대로 사용. 디버그 정보는 warnings 에 짧게 노출.
        if llm_result.model:
            warnings.append(
                f"면장 LLM 추출 성공 — model={llm_result.model}, "
                f"건수={len(llm_result.rows)}, cost≈${llm_result.cost_usd:.4f}"
            )
        return CustomsParseResult(rows=llm_result.rows, warnings=warnings)
    # LLM 0건 → fallback. 사유는 운영 진단을 위해 그대로 노출.
    warnings.extend(llm_result.warnings)

    # ---- 2차: pdfplumber 표 추출 ----------------------------------------
    for table in _extract_pdf_tables(file_bytes):
        detected = _find_header_row(table)
        if detected is None:
            continue
        header_idx, col_map = detected
        rows = _rows_to_customs(table, header_idx, col_map)
        if rows:
            warnings.append(
                "면장 PDF 표 추출(헤더 매핑)으로 파싱했습니다. "
                "LLM 추출이 실패해 fallback 경로가 사용됐습니다."
            )
            return CustomsParseResult(rows=rows, warnings=warnings)

    # ---- 3차: 본문 텍스트 라인 정규식 fallback --------------------------
    rows = _parse_pdf_lines(text)
    if not rows:
        warnings.append(
            "면장 PDF 에서 신고번호·신고가를 인식하지 못했습니다 "
            "(LLM · 표 · 라인 정규식 모두 실패). "
            "PDF 가 면장 양식이 맞는지, 텍스트 레이어가 살아있는지 확인하세요."
        )
    else:
        warnings.append(
            "면장 PDF 를 라인 기반(정규식) fallback 으로 파싱했습니다. "
            "신고번호·신고가 외 컬럼(품명/재고/BL)은 누락될 수 있으니 결과를 확인하세요."
        )
    return CustomsParseResult(rows=rows, warnings=warnings)


def parse_customs_file(file_bytes: bytes, filename: str) -> CustomsParseResult:
    """확장자로 엑셀/PDF 파서를 분기하는 진입점.

    - .xlsx / .xlsm → parse_customs_sheet
    - .pdf          → parse_customs_pdf
    - 그 외          → rows=[] + 경고 (라우터가 확장자를 1차 검증하지만 방어적 처리).
    """
    ext = Path(filename).suffix.lower()
    if ext in (".xlsx", ".xlsm"):
        return parse_customs_sheet(file_bytes)
    if ext == ".pdf":
        return parse_customs_pdf(file_bytes, source_file_name=filename)
    return CustomsParseResult(
        rows=[],
        warnings=[f"지원하지 않는 파일 형식입니다: {ext} (.xlsx/.xlsm/.pdf 만 가능)"],
    )
