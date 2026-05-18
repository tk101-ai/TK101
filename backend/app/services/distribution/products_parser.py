"""명품재고대장 엑셀 파서 (T9 Phase B-1).

대상 시트: "명품재고대장" (449r × 27c) — 41MB 큰 파일의 핵심 시트.
헤더 R6 / 데이터 R7~ (관찰 기준).

추출 컬럼 (DB → 엑셀 헤더):
- brand                ← 브랜드명
- product_name_en      ← IT물품내역(영문)
- product_code         ← 제품 코드 번호
- category             ← 카테고리 (Bag/Belts/Ring/Scarf)
- purchase_qty         ← 매입 수량
- domestic_stock_qty   ← 국내 재고 수량 (한국 인천 창고)
- supply_price         ← 공급가액
- vat                  ← 부가세
- purchase_price       ← 매입금액
- approval_number      ← 승인번호
- purchase_date        ← 매입 일시

빈 셀 / 헤더 행 / 합계 행 등은 skip.
"""
from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import openpyxl

logger = logging.getLogger(__name__)

# 헤더 → DB 컬럼 매핑. 키는 정규화된 엑셀 헤더 텍스트.
_HEADER_MAP: dict[str, str] = {
    "브랜드명": "brand",
    "IT물품내역(영문)": "product_name_en",
    "IT물품내역": "product_name_en",
    "제품 코드 번호": "product_code",
    "제품코드번호": "product_code",
    "카테고리": "category",
    "매입 수량": "purchase_qty",
    "매입수량": "purchase_qty",
    "국내 재고 수량": "domestic_stock_qty",
    "공급가액": "supply_price",
    "부가세": "vat",
    "매입금액": "purchase_price",
    "승인번호": "approval_number",
    "매입 일시": "purchase_date",
    "매입일시": "purchase_date",
}


@dataclass
class ProductRow:
    """파싱된 제품 1행."""

    brand: str
    product_name_en: str | None = None
    product_code: str | None = None
    category: str | None = None
    purchase_qty: int | None = None
    domestic_stock_qty: int | None = None
    supply_price: Decimal | None = None
    vat: Decimal | None = None
    purchase_price: Decimal | None = None
    approval_number: str | None = None
    purchase_date: date | None = None
    raw_row: dict[str, Any] = field(default_factory=dict)


@dataclass
class ProductsParseResult:
    rows: list[ProductRow]
    warnings: list[str]


def _normalize_header(raw: object) -> str | None:
    if raw is None:
        return None
    s = re.sub(r"\s+", " ", str(raw)).strip()
    # 줄바꿈으로 인한 부연설명 제거
    s = re.sub(r"\n.*", "", s).strip()
    return s or None


def _to_int(value: object) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(Decimal(str(value)))
    except (InvalidOperation, ValueError):
        return None


def _to_decimal(value: object) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def _to_date(value: object) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, date):
        return value
    if isinstance(value, datetime):
        return value.date()
    s = str(value).strip()
    # "2025.12.09", "2025-12-09", "2025-12-01 15:09:00" 등 처리
    for fmt in ("%Y.%m.%d", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _find_header_row(
    rows: list[tuple[object, ...]], max_scan: int = 20
) -> tuple[int, dict[str, int]] | None:
    """헤더 행 인덱스 + {정규화된 header: column_index} 매핑 반환.

    "브랜드명" 셀이 발견된 행을 헤더로 간주.
    """
    for idx, row in enumerate(rows[:max_scan]):
        col_map: dict[str, int] = {}
        for col_idx, cell in enumerate(row):
            header = _normalize_header(cell)
            if header and header in _HEADER_MAP:
                col_map[_HEADER_MAP[header]] = col_idx
        # 최소한 brand + product_name_en 또는 product_code 있어야 헤더 인정.
        if "brand" in col_map and (
            "product_name_en" in col_map or "product_code" in col_map
        ):
            return idx, col_map
    return None


def parse_products_sheet(
    path: Path | str,
    *,
    sheet_name: str = "명품재고대장",
) -> ProductsParseResult:
    """명품재고대장 시트 파싱."""
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        return ProductsParseResult(
            rows=[],
            warnings=[f"시트 '{sheet_name}' 없음. 가용: {wb.sheetnames}"],
        )
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    detected = _find_header_row(rows)
    if detected is None:
        return ProductsParseResult(
            rows=[], warnings=["헤더 행을 찾지 못했습니다 (브랜드명 등)"]
        )
    header_idx, col_map = detected

    parsed: list[ProductRow] = []
    warnings: list[str] = []

    for r_idx, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        if all(c is None or str(c).strip() == "" for c in row):
            continue

        brand_raw = row[col_map["brand"]] if col_map["brand"] < len(row) else None
        if brand_raw is None or str(brand_raw).strip() == "":
            continue  # 브랜드 없으면 skip (합계/소계 행 등)
        brand = str(brand_raw).strip()
        # "항목" 같은 헤더 잔재 skip
        if brand in {"항목", "브랜드명"}:
            continue

        def get(col_key: str) -> object:
            idx = col_map.get(col_key)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        product = ProductRow(
            brand=brand,
            product_name_en=(str(get("product_name_en")).strip() if get("product_name_en") else None),
            product_code=(str(get("product_code")).strip() if get("product_code") else None),
            category=(str(get("category")).strip() if get("category") else None),
            purchase_qty=_to_int(get("purchase_qty")),
            domestic_stock_qty=_to_int(get("domestic_stock_qty")),
            supply_price=_to_decimal(get("supply_price")),
            vat=_to_decimal(get("vat")),
            purchase_price=_to_decimal(get("purchase_price")),
            approval_number=(str(get("approval_number")).strip() if get("approval_number") else None),
            purchase_date=_to_date(get("purchase_date")),
            raw_row={
                k: (str(row[v]) if v < len(row) and row[v] is not None else None)
                for k, v in col_map.items()
            },
        )
        parsed.append(product)

    return ProductsParseResult(rows=parsed, warnings=warnings)
