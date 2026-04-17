"""Bank statement Excel parser for 6 Korean banks.

Supported banks: KB국민, IBK기업, NH농협, 신한, 우리, 하나
"""
from __future__ import annotations

import io
import logging
from datetime import date
from decimal import Decimal, InvalidOperation
from typing import Callable

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

BANK_ALIASES: dict[str, str] = {
    "kb": "kb", "국민": "kb", "KB국민": "kb", "KB": "kb",
    "ibk": "ibk", "기업": "ibk", "기업은행": "ibk", "IBK": "ibk",
    "nh": "nh", "농협": "nh", "NH": "nh",
    "shinhan": "shinhan", "신한": "shinhan",
    "woori": "woori", "우리": "woori",
    "hana": "hana", "하나": "hana",
}


def _normalise_bank(name: str) -> str:
    key = name.strip()
    result = BANK_ALIASES.get(key)
    if result is None:
        raise ValueError(
            f"알 수 없는 은행: '{name}'. "
            f"지원: {', '.join(sorted(set(BANK_ALIASES.values())))}"
        )
    return result


_SUMMARY_KEYWORDS = ("합계", "총", "소계")


def _is_summary_row(row: tuple) -> bool:
    for cell in row:
        if cell is None:
            continue
        text = str(cell).strip()
        if any(text.startswith(kw) for kw in _SUMMARY_KEYWORDS):
            return True
        if text.startswith("총 ") and text.endswith("건"):
            return True
    return False


def _to_decimal(value: object) -> Decimal | None:
    if value is None:
        return None
    try:
        text = str(value).replace(",", "").strip()
        if not text:
            return None
        d = Decimal(text)
        return d if d != 0 else None
    except (InvalidOperation, ValueError):
        return None


def _strip_or_none(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def _parse_date(text: str) -> date | None:
    try:
        cleaned = text.replace(".", "-").replace("/", "-").strip()
        return date.fromisoformat(cleaned[:10])
    except (ValueError, IndexError):
        return None


def _cell_date(value: object) -> date | None:
    if value is None:
        return None
    if isinstance(value, date):
        return value
    return _parse_date(str(value))


def _build_col_index(headers: list[str]) -> dict[str, int]:
    return {h: i for i, h in enumerate(headers) if h}


def _check_foreign_currency(headers: list[str], sheet_title: str) -> None:
    for h in headers:
        if "통화" in h:
            raise ValueError("외화 계좌는 아직 지원하지 않습니다")
    title_lower = sheet_title.lower()
    if "외환" in title_lower or "외화" in title_lower:
        raise ValueError("외화 계좌는 아직 지원하지 않습니다")


def _parse_kb(rows: list[tuple], headers: list[str]) -> list[dict]:
    results: list[dict] = []
    for row in rows:
        if not row or all(c is None for c in row):
            continue
        if _is_summary_row(row):
            continue
        txn_date = _cell_date(row[1])  # col B
        if txn_date is None:
            continue
        withdrawal = _to_decimal(row[3])  # col D
        deposit = _to_decimal(row[4])     # col E
        if deposit:
            amount, txn_type = deposit, "deposit"
        elif withdrawal:
            amount, txn_type = withdrawal, "withdrawal"
        else:
            continue
        results.append({
            "transaction_date": txn_date,
            "amount": amount,
            "balance": _to_decimal(row[5]),  # col F
            "counterpart_name": _strip_or_none(row[2]),  # col C
            "description": _strip_or_none(row[8]) if len(row) > 8 else None,  # col I
            "transaction_type": txn_type,
        })
    return results


def _parse_ibk(rows: list[tuple], headers: list[str]) -> list[dict]:
    results: list[dict] = []
    for row in rows:
        if not row or all(c is None for c in row):
            continue
        if _is_summary_row(row):
            continue
        txn_date = _cell_date(row[1])  # col B
        if txn_date is None:
            continue
        withdrawal = _to_decimal(row[2])  # col C
        deposit = _to_decimal(row[3])     # col D
        if deposit:
            amount, txn_type = deposit, "deposit"
        elif withdrawal:
            amount, txn_type = withdrawal, "withdrawal"
        else:
            continue
        # Counterpart: col F first, fallback col M
        counterpart = _strip_or_none(row[5]) if len(row) > 5 else None
        if not counterpart and len(row) > 12:
            counterpart = _strip_or_none(row[12])
        results.append({
            "transaction_date": txn_date,
            "amount": amount,
            "balance": _to_decimal(row[4]),  # col E
            "counterpart_name": counterpart,
            "description": _strip_or_none(row[9]) if len(row) > 9 else None,  # col J
            "transaction_type": txn_type,
        })
    return results


def _parse_nh(rows: list[tuple], headers: list[str]) -> list[dict]:
    results: list[dict] = []
    for row in rows:
        if not row or all(c is None for c in row):
            continue
        if _is_summary_row(row):
            continue
        txn_date = _cell_date(row[1])  # col B
        if txn_date is None:
            continue
        withdrawal = _to_decimal(row[2])  # col C
        deposit = _to_decimal(row[3])     # col D
        # NH: empty amounts are None (not 0), so both can be None
        if deposit:
            amount, txn_type = deposit, "deposit"
        elif withdrawal:
            amount, txn_type = withdrawal, "withdrawal"
        else:
            continue
        results.append({
            "transaction_date": txn_date,
            "amount": amount,
            "balance": _to_decimal(row[4]),  # col E
            "counterpart_name": _strip_or_none(row[6]) if len(row) > 6 else None,  # col G
            "description": _strip_or_none(row[5]) if len(row) > 5 else None,  # col F
            "transaction_type": txn_type,
        })
    return results


def _parse_shinhan(rows: list[tuple], headers: list[str]) -> list[dict]:
    results: list[dict] = []
    for row in rows:
        if not row or all(c is None for c in row):
            continue
        if _is_summary_row(row):
            continue
        txn_date = _cell_date(row[2])  # col C
        if txn_date is None:
            continue
        deposit = _to_decimal(row[4])     # col E
        withdrawal = _to_decimal(row[5])  # col F
        if deposit:
            amount, txn_type = deposit, "deposit"
        elif withdrawal:
            amount, txn_type = withdrawal, "withdrawal"
        else:
            continue
        results.append({
            "transaction_date": txn_date,
            "amount": amount,
            "balance": _to_decimal(row[7]) if len(row) > 7 else None,  # col H
            "counterpart_name": _strip_or_none(row[6]) if len(row) > 6 else None,  # col G
            "description": _strip_or_none(row[3]),  # col D
            "transaction_type": txn_type,
        })
    return results


def _parse_woori(rows: list[tuple], headers: list[str]) -> list[dict]:
    results: list[dict] = []
    for row in rows:
        if not row or all(c is None for c in row):
            continue
        if _is_summary_row(row):
            continue
        txn_date = _cell_date(row[1])  # col B
        if txn_date is None:
            continue
        # REVERSED: col E = 지급(withdrawal), col F = 입금(deposit)
        withdrawal = _to_decimal(row[4])  # col E 지급
        deposit = _to_decimal(row[5])     # col F 입금
        if deposit:
            amount, txn_type = deposit, "deposit"
        elif withdrawal:
            amount, txn_type = withdrawal, "withdrawal"
        else:
            continue
        results.append({
            "transaction_date": txn_date,
            "amount": amount,
            "balance": _to_decimal(row[6]) if len(row) > 6 else None,  # col G
            "counterpart_name": _strip_or_none(row[3]),  # col D
            "description": _strip_or_none(row[2]),  # col C
            "transaction_type": txn_type,
        })
    return results


def _find_col_by_header(headers: list[str], keyword: str) -> int | None:
    for i, h in enumerate(headers):
        if keyword in h:
            return i
    return None


def _parse_hana(rows: list[tuple], headers: list[str]) -> list[dict]:
    # Hana columns can shift if 추가메모 column exists
    deposit_col = _find_col_by_header(headers, "입금")
    withdrawal_col = _find_col_by_header(headers, "출금")
    balance_col = _find_col_by_header(headers, "거래후잔액")

    if deposit_col is None or withdrawal_col is None:
        raise ValueError("하나은행: 입금/출금 컬럼을 찾을 수 없습니다")

    # Counterpart: look for 의뢰인/수취인
    counterpart_col = _find_col_by_header(headers, "의뢰인") or _find_col_by_header(headers, "수취인")

    results: list[dict] = []
    for row in rows:
        if not row or all(c is None for c in row):
            continue
        if _is_summary_row(row):
            continue
        txn_date = _cell_date(row[0])  # col A
        if txn_date is None:
            continue
        deposit = _to_decimal(row[deposit_col]) if deposit_col < len(row) else None
        withdrawal = _to_decimal(row[withdrawal_col]) if withdrawal_col < len(row) else None
        if deposit:
            amount, txn_type = deposit, "deposit"
        elif withdrawal:
            amount, txn_type = withdrawal, "withdrawal"
        else:
            continue
        balance = None
        if balance_col is not None and balance_col < len(row):
            balance = _to_decimal(row[balance_col])
        counterpart = None
        if counterpart_col is not None and counterpart_col < len(row):
            counterpart = _strip_or_none(row[counterpart_col])
        results.append({
            "transaction_date": txn_date,
            "amount": amount,
            "balance": balance,
            "counterpart_name": counterpart,
            "description": _strip_or_none(row[1]),  # col B 적요
            "transaction_type": txn_type,
        })
    return results


_BANK_CONFIGS: dict[str, tuple[int, bool, Callable]] = {
    #  bank_key: (header_row_0indexed, multi_sheet, parser_fn)
    "kb":      (6,  False, _parse_kb),
    "ibk":     (2,  True,  _parse_ibk),
    "nh":      (9,  False, _parse_nh),
    "shinhan": (3,  True,  _parse_shinhan),
    "woori":   (3,  False, _parse_woori),
    "hana":    (6,  True,  _parse_hana),
}

# Auto-detect markers: (header_keyword, bank_key)
_DETECT_MARKERS: list[tuple[str, str]] = [
    ("보낸분/받는분",     "kb"),
    ("상대계좌번호",      "ibk"),
    ("CMS코드",          "ibk"),
    ("거래기록사항",      "nh"),
    ("전체선택",          "shinhan"),
    ("지급(원)",          "woori"),
    ("의뢰인/수취인",    "hana"),
    ("거래특이사항",      "hana"),
]


def _detect_bank(wb) -> str:  # type: ignore[no-untyped-def]
    for ws in wb.worksheets:
        # Scan first 15 rows for header markers
        for row in ws.iter_rows(min_row=1, max_row=15, values_only=True):
            row_text = " ".join(str(c or "") for c in row)
            for marker, bank_key in _DETECT_MARKERS:
                if marker in row_text:
                    logger.info("Auto-detected bank: %s (marker: %s)", bank_key, marker)
                    return bank_key
    raise ValueError(
        "은행 형식을 자동 감지할 수 없습니다. "
        "bank_name 파라미터를 지정해 주세요."
    )


def _parse_sheet(
    ws,  # type: ignore[no-untyped-def]
    header_row: int,
    parser_fn: Callable,
) -> list[dict]:
    all_rows = list(ws.iter_rows(values_only=True))
    if len(all_rows) <= header_row + 1:
        return []

    headers = [str(c or "").strip() for c in all_rows[header_row]]

    # Foreign currency check
    _check_foreign_currency(headers, ws.title or "")

    data_rows = all_rows[header_row + 1:]
    return parser_fn(data_rows, headers)


def parse_bank_excel(
    file_bytes: bytes | io.BytesIO,
    bank_name: str | None = None,
) -> list[dict]:
    """Parse a Korean bank statement Excel file.

    Returns list of transaction dicts. If *bank_name* is None,
    auto-detects from header keywords.
    """
    if isinstance(file_bytes, (bytes, bytearray)):
        file_bytes = io.BytesIO(file_bytes)

    wb = load_workbook(filename=file_bytes, read_only=True, data_only=True)

    try:
        if bank_name is not None:
            bank_key = _normalise_bank(bank_name)
        else:
            bank_key = _detect_bank(wb)

        header_row, multi_sheet, parser_fn = _BANK_CONFIGS[bank_key]
        logger.info("Parsing with bank format: %s", bank_key)

        transactions: list[dict] = []

        if multi_sheet:
            for ws in wb.worksheets:
                sheet_txns = _parse_sheet(ws, header_row, parser_fn)
                transactions.extend(sheet_txns)
                if sheet_txns:
                    logger.info(
                        "Sheet '%s': %d transactions", ws.title, len(sheet_txns)
                    )
        else:
            ws = wb.active
            transactions = _parse_sheet(ws, header_row, parser_fn)

        logger.info("Total transactions parsed: %d", len(transactions))
        return transactions
    finally:
        wb.close()
