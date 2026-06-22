"""SNS 데이터 → .xlsx 내보내기 헬퍼 (1단계).

각 함수는 행(row) 리스트를 받아 openpyxl ``Workbook`` 을 만들고 그 바이트(io.BytesIO)를
돌려준다. DB 접근은 일절 하지 않는다 — 라우터가 기존 쿼리로 행을 만들어 넘겨준다.

다중 클라이언트 안전:
    모든 시트는 계정(채널) 단위로 행을 돌며, 브랜드(client)/플랫폼/어권/핸들을 식별
    컬럼으로 항상 포함한다. 서울시 등 특정 발주처/채널을 하드코딩하지 않는다 —
    계정 쿼리에 행이 추가되면 코드 변경 없이 그대로 내보내기에 반영된다.
"""
from __future__ import annotations

import io
from typing import Any, Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# 어권/플랫폼 코드 → 사람이 읽는 한글 라벨. 미지의 값은 원본 코드를 그대로 쓴다
# (새 어권/플랫폼이 추가돼도 깨지지 않도록 — 데이터 주도).
LANGUAGE_LABELS: dict[str, str] = {"en": "영문", "zh": "중간체", "ja": "일문"}
PLATFORM_LABELS: dict[str, str] = {
    "facebook": "페이스북",
    "instagram": "인스타그램",
    "twitter": "트위터",
    "youtube": "유튜브",
    "weibo": "웨이보",
}

_HEADER_FONT = Font(bold=True)
_TOTAL_FONT = Font(bold=True)


def _label(mapping: dict[str, str], value: Any) -> str:
    """코드를 한글 라벨로. 없으면 원본 문자열, None 이면 빈 칸."""
    if value is None:
        return ""
    return mapping.get(str(value), str(value))


def _identity_cells(row: Any) -> list[Any]:
    """[브랜드(client), 플랫폼, 어권, 핸들] — 모든 시트 공통 식별 컬럼."""
    return [
        getattr(row, "client", None) or "",
        _label(PLATFORM_LABELS, getattr(row, "platform", None)),
        _label(LANGUAGE_LABELS, getattr(row, "language", None)),
        getattr(row, "handle", None) or "",
    ]


def _autosize(ws: Any, headers: Sequence[str], sample_widths: Sequence[int] | None = None) -> None:
    """간단한 컬럼 너비 — 헤더 길이 기반(+여유). 한글은 폭이 넓어 1.6배 가중."""
    for idx, header in enumerate(headers, start=1):
        base = max(len(str(header)) * 1.6, 8)
        if sample_widths and idx - 1 < len(sample_widths):
            base = max(base, sample_widths[idx - 1])
        ws.column_dimensions[get_column_letter(idx)].width = min(base + 2, 60)


def _write_header(ws: Any, headers: Sequence[str]) -> None:
    ws.append(list(headers))
    for cell in ws[1]:
        cell.font = _HEADER_FONT


def _finalize(wb: Workbook) -> io.BytesIO:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# 주차 컬럼 5개 + 합계. 콘텐츠 현황 시트가 쓴다.
_WEEK_HEADERS = ["1주차", "2주차", "3주차", "4주차", "5주차"]


def build_content_status_workbook(rows: Iterable[Any]) -> io.BytesIO:
    """콘텐츠 현황(주차별 게재건수) — 계정당 1행 + 맨 아래 전체 합계.

    행(row)은 ``WeeklyPostCountRow`` 류로 client/platform/language/handle 과
    week1..week5, total 속성을 갖는다(stats_weekly_posts 산출물).
    """
    headers = ["브랜드", "플랫폼", "어권", "핸들", *_WEEK_HEADERS, "합계"]
    wb = Workbook()
    ws = wb.active
    ws.title = "콘텐츠현황"
    _write_header(ws, headers)

    totals = [0, 0, 0, 0, 0]
    grand_total = 0
    for row in rows:
        weeks = [int(getattr(row, f"week{i}", 0) or 0) for i in range(1, 6)]
        total = int(getattr(row, "total", 0) or 0)
        for i in range(5):
            totals[i] += weeks[i]
        grand_total += total
        ws.append([*_identity_cells(row), *weeks, total])

    # 전체 합계 행.
    total_row_idx = ws.max_row + 1
    ws.append(["전체 합계", "", "", "", *totals, grand_total])
    for cell in ws[total_row_idx]:
        cell.font = _TOTAL_FONT

    _autosize(ws, headers)
    return _finalize(wb)


def build_snapshots_workbook(rows: Iterable[Any]) -> io.BytesIO:
    """주간 팔로워 — 계정당 1행, 1~5주차 팔로워 수.

    행(row)은 client/platform/language/handle 과 week1..week5(팔로워, None 가능)를 갖는다.
    """
    headers = ["브랜드", "플랫폼", "어권", "핸들", *_WEEK_HEADERS]
    wb = Workbook()
    ws = wb.active
    ws.title = "주간팔로워"
    _write_header(ws, headers)

    for row in rows:
        weeks = [getattr(row, f"week{i}", None) for i in range(1, 6)]
        # None(미입력)은 빈 칸으로 둔다.
        cells = ["" if w is None else int(w) for w in weeks]
        ws.append([*_identity_cells(row), *cells])

    _autosize(ws, headers)
    return _finalize(wb)


def build_posts_workbook(rows: Iterable[Any]) -> io.BytesIO:
    """게시물 목록 — 게시물당 1행. 계정 식별 컬럼 + 게시물 필드.

    행(row)은 client/platform/language/handle(계정) 과
    posted_at/title/content_type/producer/view_count/reach_count/comment_count/
    like_count/share_count/total_engagement/url(게시물) 속성을 갖는다.
    """
    headers = [
        "브랜드",
        "플랫폼",
        "어권",
        "핸들",
        "게재일",
        "제목",
        "형태",
        "제작",
        "조회수",
        "도달",
        "댓글",
        "좋아요",
        "공유",
        "토탈인게이지먼트",
        "링크",
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "게시물"
    _write_header(ws, headers)

    for row in rows:
        posted_at = getattr(row, "posted_at", None)
        ws.append(
            [
                *_identity_cells(row),
                "" if posted_at is None else str(posted_at),
                getattr(row, "title", None) or "",
                getattr(row, "content_type", None) or "",
                getattr(row, "producer", None) or "",
                _num(getattr(row, "view_count", None)),
                _num(getattr(row, "reach_count", None)),
                _num(getattr(row, "comment_count", None)),
                _num(getattr(row, "like_count", None)),
                _num(getattr(row, "share_count", None)),
                _num(getattr(row, "total_engagement", None)),
                getattr(row, "url", None) or "",
            ]
        )

    _autosize(ws, headers)
    return _finalize(wb)


def _num(value: Any) -> Any:
    """숫자 칸 — None 이면 빈 칸, 아니면 정수로."""
    if value is None:
        return ""
    try:
        return int(value)
    except (TypeError, ValueError):
        return value


# ==================================================================
# 2단계 — 브랜드별 통합 워크북 (월간요약 + 채널별 콘텐츠)
# ==================================================================
#
# marketing1 importer 와 라운드트립(round-trip) 호환되도록, 팀의 기존 구글시트
# 구조(`[TK내부] 2026 서울시 글로벌 SNS DB 관리.xlsx`)를 역으로 재현한다.
#
# importer(`sns_importers/marketing1.py`)가 읽는 규칙에 정확히 맞춘다:
#   - 통합(월간요약) 시트: 헤더 행3.
#       * 컬럼B(2): "N월" / "영문|중문|일문" / "구분" / "합계" 마커
#       * 컬럼C(3): 플랫폼 텍스트(페이스북/인스타그램/…)
#       * 주차 팔로워 컬럼 = D(4),F(6),H(8),J(10)  (사이 칸은 WoW 증감률)
#   - 채널별 콘텐츠 시트: 헤더 행3, 헤더 텍스트로 동적 매핑(HEADER_ALIASES).
#
# importer 의 LANGUAGE_TEXT_MAP 은 zh → "중문" 이므로(이 모듈의 "중간체" 와 다름),
# 통합 시트의 어권 마커는 importer 호환 라벨을 따로 쓴다(_SUMMARY_LANGUAGE_LABELS).

# importer 의 LANGUAGE_TEXT_MAP 을 역으로 — 통합 시트 어권 마커용(라운드트립).
_SUMMARY_LANGUAGE_LABELS: dict[str, str] = {"en": "영문", "zh": "중문", "ja": "일문"}

# 통합 시트 — 주차별 (팔로워 컬럼, WoW 컬럼) 1-based 인덱스. importer 의
# WEEK_FOLLOWER_COLUMNS({1:4,2:6,3:8,4:10}) 와 정합. 5주차도 같은 패턴으로 확장.
_SUMMARY_WEEK_COLUMNS: dict[int, tuple[int, int]] = {
    1: (4, 5),
    2: (6, 7),
    3: (8, 9),
    4: (10, 11),
    5: (12, 13),
}
# 5주차 팔로워/WoW 다음 컬럼들 — 주차별 게재건수, 월 누적, 인터랙션/조회수.
_SUMMARY_UPLOAD_START = 14  # 1~5주차 게재건수: 14~18
_SUMMARY_CUMULATIVE_COL = 19  # 월 누적 게재건수
_SUMMARY_INTERACTION_COL = 20  # 인터랙션/조회수


def _summary_headers() -> list[str]:
    """통합(월간요약) 시트 행3 헤더. importer 는 통합 시트 헤더를 직접 읽지 않지만,
    사람이 보는 라벨로서 컬럼 정렬을 위해 둔다."""
    headers = ["", "구분", "채널", "1주차", "증감률"]
    for week in range(2, 6):
        headers += [f"{week}주차", "증감률"]
    headers += [f"{w}주차 게재" for w in range(1, 6)]
    headers += ["월 누적", "인터랙션/조회수"]
    return headers


def _wow_rate(curr: int | None, prev: int | None) -> Any:
    """주간 팔로워 증감률(WoW). 직전 주차 대비. 미입력/0 분모면 빈 칸."""
    if curr is None or prev is None or prev == 0:
        return ""
    return round((curr - prev) / prev, 4)


def _content_sheet_title(platform: str, language: str) -> str:
    """채널별 콘텐츠 시트명. 예: '페이스북 (영문)'. 31자 제한 + 중복 회피는 호출자."""
    p = _label(PLATFORM_LABELS, platform)
    lang = _label(_SUMMARY_LANGUAGE_LABELS, language)
    return f"{p} ({lang})"[:31]


# 채널별 콘텐츠 시트 헤더(행3). importer 의 HEADER_ALIASES 와 부분일치하도록 라벨을
# 고른다(월/번호/주차/배포일/콘텐츠 제목/콘텐츠 형태/제작/조회수/도달/댓글/좋아요/
# 공유/스크랩/리포스팅/토탈 인게이지먼트/링크/데이터 기입일자).
_CONTENT_HEADERS: list[str] = [
    "월",
    "번호",
    "주차",
    "배포일",
    "콘텐츠 제목",
    "콘텐츠 형태",
    "구분",
    "제작",
    "노출수",
    "조회수",
    "도달",
    "댓글",
    "좋아요",
    "리포스팅",
    "공유",
    "스크랩",
    "토탈 인게이지먼트",
    "링크",
    "데이터 기입일자",
]

# 콘텐츠 형태 enum → 한글(시트 표기). importer 의 CONTENT_TYPE_MAP 역방향.
_CONTENT_TYPE_LABELS: dict[str, str] = {
    "image": "이미지",
    "video": "영상",
    "short": "숏폼",
}


def _week_of_month(d: Any) -> int | None:
    """배포일 → 월중 주차(1~5). stats_weekly_posts 와 동일 공식 ((day-1)//7)+1."""
    day = getattr(d, "day", None)
    if day is None:
        return None
    return ((day - 1) // 7) + 1


def _date_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def build_full_brand_workbook(
    client: str,
    year: int,
    month: int,
    accounts: Iterable[Any],
    summary_rows: Iterable[Any],
    posts_by_account: dict[Any, Sequence[Any]],
    snapshots_by_account: dict[Any, dict[int, Any]],
) -> io.BytesIO:
    """브랜드(client) 통합 워크북 — 월간요약 1시트 + 채널별 콘텐츠 N시트.

    DB 미접근(데이터는 라우터가 넘긴다). 특정 발주처/채널을 하드코딩하지 않는다 —
    ``accounts`` 를 순회하므로 신세계 등 다른 브랜드도 자동으로 자기 워크북을 만든다.

    Args:
        client: 브랜드명(시트 식별/파일명). 빈 값이면 "전체".
        year, month: 요약 블록이 대상으로 삼는 연/월.
        accounts: 이 브랜드의 계정들. 각 항목은 id/platform/language/handle 속성.
        summary_rows: 계정별 주차 게재건수 행(stats_weekly_posts 산출물).
            account_id/week1..week5/total 속성.
        posts_by_account: {account_id: [post, ...]}. post 는 posted_at/title/
            content_type/producer/view_count/reach_count/comment_count/like_count/
            share_count/save_count/repost_count/total_engagement/url/data_recorded_at.
        snapshots_by_account: {account_id: {week_number: followers}}. 주차별 팔로워.

    Returns:
        io.BytesIO (xlsx 바이트, seek(0)).
    """
    summary_by_account = {getattr(r, "account_id", None): r for r in summary_rows}
    account_list = list(accounts)

    wb = Workbook()
    _build_summary_sheet(
        wb,
        year=year,
        month=month,
        accounts=account_list,
        summary_by_account=summary_by_account,
        snapshots_by_account=snapshots_by_account,
    )
    _build_channel_sheets(wb, account_list, posts_by_account, month=month)
    return _finalize(wb)


def _build_summary_sheet(
    wb: Workbook,
    *,
    year: int,
    month: int,
    accounts: Sequence[Any],
    summary_by_account: dict[Any, Any],
    snapshots_by_account: dict[Any, dict[int, Any]],
) -> None:
    """월간요약 시트 — importer 가 다시 읽을 수 있는 통합 시트 레이아웃.

    행1: 제목, 행2: 월 마커("N월"), 행3: 헤더. 이후 어권 그룹별로 채널 행 + 합계.
    importer 는 컬럼B 의 월/어권 마커와 컬럼C 의 플랫폼, 컬럼 4/6/8/10 의 팔로워를 읽는다.
    """
    # 시트명 "통합" — importer 의 parse_workbook 이 통합 시트로 인식해 라운드트립한다.
    ws = wb.active
    ws.title = "통합"

    ws.cell(row=1, column=2, value=f"{year}년 {month}월 SNS 월간요약")
    ws.cell(row=1, column=2).font = _HEADER_FONT
    # 행2: 월 마커(컬럼B). importer 가 "N월" 로 현재 월을 잡는다.
    ws.cell(row=2, column=2, value=f"{month}월")
    # 행3: 헤더.
    for idx, header in enumerate(_summary_headers(), start=1):
        cell = ws.cell(row=3, column=idx, value=header)
        cell.font = _HEADER_FONT

    # 어권 → 그 어권의 계정들(정렬). importer 는 어권 마커를 만나면 갱신한다.
    by_language: dict[str, list[Any]] = {}
    for acc in accounts:
        by_language.setdefault(getattr(acc, "language", None), []).append(acc)

    row = 4
    grand = _SummaryAggregate()
    for language in sorted(by_language, key=lambda x: x or ""):
        # 어권 마커 행(컬럼B). 같은 행 컬럼C 부터 첫 채널을 함께 둘 수도 있으나,
        # 명확성을 위해 어권 마커 행과 채널 행을 분리한다(importer 는 둘 다 처리).
        ws.cell(row=row, column=2, value=_label(_SUMMARY_LANGUAGE_LABELS, language))
        row += 1
        accs = sorted(
            by_language[language],
            key=lambda a: (getattr(a, "platform", "") or "", getattr(a, "handle", "") or ""),
        )
        for acc in accs:
            _write_summary_channel_row(
                ws,
                row,
                acc,
                summary_by_account.get(getattr(acc, "id", None)),
                snapshots_by_account.get(getattr(acc, "id", None), {}),
                grand,
            )
            row += 1

    # 합계 행(컬럼B "합계"). importer 는 "합계" 행을 스킵한다.
    ws.cell(row=row, column=2, value="합계").font = _TOTAL_FONT
    for week, follower_total in grand.followers.items():
        f_col, _ = _SUMMARY_WEEK_COLUMNS[week]
        if follower_total:
            c = ws.cell(row=row, column=f_col, value=follower_total)
            c.font = _TOTAL_FONT
    for week in range(1, 6):
        c = ws.cell(row=row, column=_SUMMARY_UPLOAD_START + week - 1, value=grand.uploads[week])
        c.font = _TOTAL_FONT
    ws.cell(row=row, column=_SUMMARY_CUMULATIVE_COL, value=grand.cumulative).font = _TOTAL_FONT
    ws.cell(row=row, column=_SUMMARY_INTERACTION_COL, value=grand.interaction).font = _TOTAL_FONT

    _autosize(ws, _summary_headers())


class _SummaryAggregate:
    """합계 행 누적용. 주차별 팔로워/게재건수 + 월 누적 + 인터랙션."""

    def __init__(self) -> None:
        self.followers: dict[int, int] = {w: 0 for w in range(1, 6)}
        self.uploads: dict[int, int] = {w: 0 for w in range(1, 6)}
        self.cumulative = 0
        self.interaction = 0


def _write_summary_channel_row(
    ws: Any,
    row: int,
    acc: Any,
    summary: Any,
    week_followers: dict[int, Any],
    grand: _SummaryAggregate,
) -> None:
    """채널 1행 — 컬럼C 플랫폼, 주차별 팔로워(+WoW), 주차별 게재건수, 월 누적, 인터랙션."""
    ws.cell(row=row, column=3, value=_label(PLATFORM_LABELS, getattr(acc, "platform", None)))

    prev_followers: int | None = None
    for week in range(1, 6):
        f_col, wow_col = _SUMMARY_WEEK_COLUMNS[week]
        followers = week_followers.get(week)
        f_int = None if followers is None else int(followers)
        if f_int is not None:
            ws.cell(row=row, column=f_col, value=f_int)
            grand.followers[week] += f_int
            rate = _wow_rate(f_int, prev_followers)
            if rate != "":
                ws.cell(row=row, column=wow_col, value=rate)
        prev_followers = f_int if f_int is not None else prev_followers

    # 주차별 게재건수 + 월 누적.
    cumulative = 0
    for week in range(1, 6):
        uploads = int(getattr(summary, f"week{week}", 0) or 0) if summary else 0
        ws.cell(row=row, column=_SUMMARY_UPLOAD_START + week - 1, value=uploads)
        grand.uploads[week] += uploads
        cumulative += uploads
    ws.cell(row=row, column=_SUMMARY_CUMULATIVE_COL, value=cumulative)
    grand.cumulative += cumulative

    interaction = int(getattr(summary, "interaction", 0) or 0) if summary else 0
    ws.cell(row=row, column=_SUMMARY_INTERACTION_COL, value=interaction)
    grand.interaction += interaction


def _build_channel_sheets(
    wb: Workbook,
    accounts: Sequence[Any],
    posts_by_account: dict[Any, Sequence[Any]],
    *,
    month: int,
) -> None:
    """계정마다 콘텐츠 시트 1개 — 게시물당 1행, 행3 헤더(importer 호환)."""
    used_titles: set[str] = set()
    for acc in accounts:
        base = _content_sheet_title(getattr(acc, "platform", ""), getattr(acc, "language", ""))
        title = base
        suffix = 2
        while title in used_titles:
            title = f"{base[:28]}_{suffix}"
            suffix += 1
        used_titles.add(title)

        ws = wb.create_sheet(title=title)
        # 행1: 시트명, 행2: 빈 줄, 행3: 헤더(importer 가 행3 헤더로 동적 매핑).
        ws.cell(row=1, column=1, value=title).font = _HEADER_FONT
        for idx, header in enumerate(_CONTENT_HEADERS, start=1):
            ws.cell(row=3, column=idx, value=header).font = _HEADER_FONT

        posts = posts_by_account.get(getattr(acc, "id", None), [])
        for n, post in enumerate(posts, start=1):
            _write_post_row(ws, 3 + n, n, post)

        _autosize(ws, _CONTENT_HEADERS)


def _write_post_row(ws: Any, row: int, number: int, post: Any) -> None:
    """게시물 1행 — _CONTENT_HEADERS 컬럼 순서."""
    posted_at = getattr(post, "posted_at", None)
    content_type = getattr(post, "content_type", None)
    cells = [
        getattr(posted_at, "month", "") or "",  # 월
        number,  # 번호
        _week_of_month(posted_at) or "",  # 주차
        _date_str(posted_at),  # 배포일
        getattr(post, "title", None) or "",  # 콘텐츠 제목
        _CONTENT_TYPE_LABELS.get(content_type, content_type or ""),  # 콘텐츠 형태
        "",  # 구분 (저장 안 함 — 빈 칸)
        getattr(post, "producer", None) or "",  # 제작
        _num(getattr(post, "reach_count", None)),  # 노출수 (reach 로 채움)
        _num(getattr(post, "view_count", None)),  # 조회수
        _num(getattr(post, "reach_count", None)),  # 도달
        _num(getattr(post, "comment_count", None)),  # 댓글
        _num(getattr(post, "like_count", None)),  # 좋아요
        _num(getattr(post, "repost_count", None)),  # 리포스팅
        _num(getattr(post, "share_count", None)),  # 공유
        _num(getattr(post, "save_count", None)),  # 스크랩
        _num(getattr(post, "total_engagement", None)),  # 토탈 인게이지먼트
        getattr(post, "url", None) or "",  # 링크
        _date_str(getattr(post, "data_recorded_at", None)),  # 데이터 기입일자
    ]
    for idx, value in enumerate(cells, start=1):
        ws.cell(row=row, column=idx, value=value)
