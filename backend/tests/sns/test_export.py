"""sns_export 워크북 빌더 단위 테스트 (DB 미접근, openpyxl 라운드트립)."""
import io
from datetime import date
from types import SimpleNamespace

from openpyxl import load_workbook

from app.services.sns_export import (
    build_content_status_workbook,
    build_full_brand_workbook,
    build_posts_workbook,
    build_snapshots_workbook,
)
from app.services.sns_importers import marketing1


def _load(buf):
    return load_workbook(buf).active


def test_content_status_has_identity_columns_and_total_row():
    rows = [
        SimpleNamespace(
            client="서울시", platform="facebook", language="en", handle="seoul.en",
            week1=2, week2=1, week3=0, week4=3, week5=0, total=6,
        ),
        SimpleNamespace(
            client="신세계", platform="instagram", language="zh", handle="ss.zh",
            week1=1, week2=1, week3=1, week4=1, week5=1, total=5,
        ),
    ]
    ws = _load(build_content_status_workbook(rows))
    assert [c.value for c in ws[1]] == [
        "브랜드", "플랫폼", "어권", "핸들",
        "1주차", "2주차", "3주차", "4주차", "5주차", "합계",
    ]
    # 코드값이 아니라 한글 라벨로 내보낸다 + 브랜드 컬럼이 데이터 주도(하드코딩 아님).
    assert ws["A2"].value == "서울시"
    assert ws["A3"].value == "신세계"
    assert ws["B2"].value == "페이스북"
    assert ws["C2"].value == "영문"
    # 전체 합계 행.
    last = ws[ws.max_row]
    assert last[0].value == "전체 합계"
    assert last[9].value == 11  # 6 + 5


def test_snapshots_blank_for_missing_weeks():
    rows = [
        SimpleNamespace(
            client=None, platform="youtube", language="ja", handle="yt.ja",
            week1=100, week2=None, week3=120, week4=None, week5=None,
        )
    ]
    ws = _load(build_snapshots_workbook(rows))
    assert ws.title == "주간팔로워"
    assert [c.value for c in ws[1]] == [
        "브랜드", "플랫폼", "어권", "핸들",
        "1주차", "2주차", "3주차", "4주차", "5주차",
    ]
    assert ws["E2"].value == 100
    assert ws["F2"].value in (None, "")  # 미입력 주차는 빈 칸


def test_posts_columns_and_date_stringified():
    rows = [
        SimpleNamespace(
            client="서울시", platform="facebook", language="en", handle="seoul.en",
            posted_at=date(2026, 4, 3), title="제목", content_type="video",
            producer="TK제작", view_count=500, reach_count=None, comment_count=2,
            like_count=10, share_count=1, total_engagement=13, url="https://x",
        )
    ]
    ws = _load(build_posts_workbook(rows))
    assert [c.value for c in ws[1]] == [
        "브랜드", "플랫폼", "어권", "핸들", "게재일", "제목", "형태", "제작",
        "조회수", "도달", "댓글", "좋아요", "공유", "토탈인게이지먼트", "링크",
    ]
    assert ws["E2"].value == "2026-04-03"
    assert ws["I2"].value == 500
    assert ws["J2"].value in (None, "")  # 도달 None
    assert ws["O2"].value == "https://x"


def test_unknown_platform_falls_back_to_raw_value():
    # 새 플랫폼/어권이 추가돼도 깨지지 않고 원본 코드를 그대로 쓴다.
    rows = [
        SimpleNamespace(
            client="X", platform="tiktok", language="ko", handle="h",
            week1=0, week2=0, week3=0, week4=0, week5=0, total=0,
        )
    ]
    ws = _load(build_content_status_workbook(rows))
    assert ws["B2"].value == "tiktok"
    assert ws["C2"].value == "ko"


# ---------------- 2단계: 브랜드 통합 워크북 ----------------


def _fake_brand_dataset():
    """작은 가짜 브랜드 데이터(DB 미접근)."""
    fb = SimpleNamespace(id="acc-fb", platform="facebook", language="en", handle="seoul.en", client="서울시")
    yt = SimpleNamespace(id="acc-yt", platform="youtube", language="en", handle="seoul.yt", client="서울시")
    accounts = [fb, yt]
    summary_rows = [
        SimpleNamespace(account_id="acc-fb", week1=2, week2=1, week3=0, week4=1, week5=0, total=4, interaction=120),
        SimpleNamespace(account_id="acc-yt", week1=1, week2=0, week3=1, week4=0, week5=0, total=2, interaction=80),
    ]
    posts_by_account = {
        "acc-fb": [
            SimpleNamespace(
                posted_at=date(2026, 4, 3), title="FB 포스트1", content_type="image", producer="TK제작",
                view_count=None, reach_count=300, comment_count=5, like_count=20, share_count=2,
                save_count=1, repost_count=None, total_engagement=27, url="https://fb/1",
                data_recorded_at=date(2026, 4, 4),
            ),
        ],
        "acc-yt": [
            SimpleNamespace(
                posted_at=date(2026, 4, 10), title="YT 영상1", content_type="video", producer="서울제작",
                view_count=5000, reach_count=None, comment_count=10, like_count=100, share_count=None,
                save_count=None, repost_count=None, total_engagement=110, url="https://yt/1",
                data_recorded_at=None,
            ),
        ],
    }
    snapshots_by_account = {
        "acc-fb": {1: 1000, 2: 1010, 3: 1010, 4: 1025},
        "acc-yt": {1: 500, 2: 520},
    }
    return accounts, summary_rows, posts_by_account, snapshots_by_account


def test_full_brand_workbook_sheets_and_channel_columns():
    accounts, summary_rows, posts_by_account, snapshots_by_account = _fake_brand_dataset()
    buf = build_full_brand_workbook(
        client="서울시", year=2026, month=4, accounts=accounts,
        summary_rows=summary_rows, posts_by_account=posts_by_account,
        snapshots_by_account=snapshots_by_account,
    )
    wb = load_workbook(buf)
    # 월간요약 시트(importer 가 통합으로 인식) + 채널별 시트.
    assert "통합" in wb.sheetnames
    assert "페이스북 (영문)" in wb.sheetnames
    assert "유튜브 (영문)" in wb.sheetnames

    # 채널 시트 헤더(행3) 가 importer alias 와 정렬되는지.
    ws_fb = wb["페이스북 (영문)"]
    header = [c.value for c in ws_fb[3]]
    for required in ("월", "주차", "배포일", "콘텐츠 제목", "콘텐츠 형태", "제작", "조회수", "도달", "링크"):
        assert required in header, f"채널 시트 헤더에 '{required}' 없음"
    # 첫 게시물 행(행4): 배포일/제목.
    posted_col = header.index("배포일") + 1
    title_col = header.index("콘텐츠 제목") + 1
    assert ws_fb.cell(row=4, column=posted_col).value == "2026-04-03"
    assert ws_fb.cell(row=4, column=title_col).value == "FB 포스트1"


def test_full_brand_workbook_summary_reimportable():
    """내보낸 통합(월간요약) 시트를 marketing1.parse_workbook 이 다시 읽는다(라운드트립)."""
    accounts, summary_rows, posts_by_account, snapshots_by_account = _fake_brand_dataset()
    buf = build_full_brand_workbook(
        client="서울시", year=2026, month=4, accounts=accounts,
        summary_rows=summary_rows, posts_by_account=posts_by_account,
        snapshots_by_account=snapshots_by_account,
    )
    parsed = marketing1.parse_workbook(buf.getvalue(), default_year=2026)

    # 통합 시트에서 계정(영문 facebook/youtube)이 복원된다.
    parsed_keys = {(a["platform"], a["language"]) for a in parsed["accounts"]}
    assert ("facebook", "en") in parsed_keys
    assert ("youtube", "en") in parsed_keys

    # 주차 팔로워 스냅샷이 importer 가 읽는 컬럼(4/6/8/10)에서 복원된다.
    fb_snaps = {
        (s["week_number"]): s["followers"]
        for s in parsed["snapshots"]
        if s["platform"] == "facebook" and s["language"] == "en" and s["month"] == 4
    }
    assert fb_snaps.get(1) == 1000
    assert fb_snaps.get(2) == 1010
    assert fb_snaps.get(4) == 1025


def test_full_brand_workbook_is_brand_agnostic():
    """다른 브랜드(신세계)도 자기 워크북을 만든다 — 서울시 하드코딩 없음."""
    acc = SimpleNamespace(id="x", platform="instagram", language="ja", handle="ss.ja", client="신세계")
    buf = build_full_brand_workbook(
        client="신세계", year=2026, month=5, accounts=[acc],
        summary_rows=[], posts_by_account={}, snapshots_by_account={},
    )
    wb = load_workbook(buf)
    assert "통합" in wb.sheetnames
    assert "인스타그램 (일문)" in wb.sheetnames
